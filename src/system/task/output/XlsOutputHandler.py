# core
import os

# contrib
import win32com.client
from openpyxl.workbook import Workbook

# MincePy
from AbstractOutputHandler import AbstractOutputHandler
from dataimport.Cell import Cell
from database.query.QueryError import QueryError

class XlsOutputHandler(AbstractOutputHandler):
    '''
    Writes query output to an Excel spreadsheet (2010/.xlsx format).
    
    :param outputPath: the file to write to; if it doesn't exist, it will be
        created
    :type outputPath: str
    :param worksheet: the name of the worksheet to write query output to
    :type worksheet: str
    :param cell: optional - the cell to begin inserting the query output at
    :type cell: str
    :param header: optional - output a header row with the column names
    :type header: boolean
    '''
    
    def __init__(self, outputPath, worksheet, cell="A1", header=False):
        self.__outputPath = outputPath
        self.__worksheet = worksheet
        self.__cell = cell
        self.__header = header


    def __writeNew(self, output):
        wb = Workbook()
        wb.save(self.__outputPath)
        
        if self.__worksheet in wb.sheetnames:
            ws = wb.get_sheet_by_name(self.__worksheet)
        else:
            ws = wb.create_sheet(title=self.__worksheet)

        if self.__header:
            ws.append(output.getColumns())
            
        for row in output:
            ws.append((value for value in row))

        wb.save(self.__outputPath)
        
        
    def __writeExisting(self, output):
        excel = None
        wb = None
           
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            wb = excel.Workbooks.Open(self.__outputPath)
            ws = wb.Sheets(self.__worksheet)
            startCell = Cell.FromString(self.__cell)
            currentRowIndex = startCell.getRowIndex() + 1
            firstColIndex = startCell.getColIndex() + 1
            
            if self.__header:
                colNames = output.getColumns()
                cells = ws.Range(
                    ws.Cells(currentRowIndex, firstColIndex),
                    ws.Cells(currentRowIndex, firstColIndex + len(colNames) - 1))
                cells.Value = colNames
                currentRowIndex += 1
            
            for row in output:
                cells = ws.Range(
                    ws.Cells(currentRowIndex, firstColIndex),
                    ws.Cells(currentRowIndex, firstColIndex + len(row) - 1))
                cells.Value = row
                currentRowIndex += 1

            excel.CalculateFull()
        except Exception as e:
            raise QueryError("Error in {}: {}".format(self.__outputPath, e))
        finally:
            if wb:
                wb.Close(True)
            if excel:
                excel.Quit()

    
    def handle(self, dbTitle, output):
        '''
        See :meth:`.AbstractOutputHandler.handle`.
        '''
        if os.path.exists(self.__outputPath):
            self.__writeExisting(output)
        else:
            self.__writeNew(output)
